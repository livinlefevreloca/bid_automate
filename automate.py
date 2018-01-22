#!python3
#automate_bid.py - automates the process of crating all the infrastructure to start a balancing bid'

import openpyxl, shutil, os

#Paths that will be used 
BID_SHEET_PATH = "//COMMSHARE2/Shared/aerotb$/0ABSMAIN/AERO_QUOTE_NUMBERS"
PROJECTS_PATH = "//COMMSHARE2/Shared/aerotb$/0ABSMAIN/1Projects"
FOLDER_TEMPLATE_PATH = "//COMMSHARE2/Shared/aerotb$/0ABSMAIN/0JOB_FOLDER_TEMPLATES/SU&CxA&TAB_Template"
PROPOSAL_PATH = "//COMMSHARE2/Shared/aerotb$/AeroServer/Lefevre/P-YYMMDD-PROJECT_NAME-CUSTOMER-(AL)"

def get_bid_data():
    #pull the data needed to create the folder and fill out the proposal summary sheet
    bid_sheet = openpyxl.load_workbook('AERO_QUOTE_NUMBERS')
    bids = bid_sheet.sheet['2018']
    last_entry = bids.get_highest_row()
    data = []
    for i in range(1,7):
        data.append(bids.cell(row=last_entry, col=i).value) 
    bid_sheet.save(BID_SHEET_PATH)
    return data
    
def create_new_bid_folder(project_name, sub_projet_name):
    #create new folder in the proper letter folder and make the folder with the project name in that folder
    folder = project_name[0]
    if str.isdigit(folder):
        folder = '1Numbers'
    else:
        folder = str.upper(folder).split(' ')
    folder_path = os.path.join('PROJECTS_PATH', folder)
    p_name = '_'.join(project_name.split(' '))
    sp_name = '_'.join(sub_projet_name.split(' '))
    project_path = os.path.join(folder_path, p_name)
    sub_project_path = os.path.join(project_path, sp_name)
    try:
        os.mkdir(project_path)
    except OSError as e:
        #if the folder exisits use the sub project name within the existing project folder
        if os.path.isdir(folder_path):
            print('folder already exisits using creating address sub projet folder')
            os.mkdir(sub_project_path)
            project_path = sub_project_path
        else:
            print('Something went wrong please see error and try again')
            print(e)
    return project_path

def copy_folder_temp(proj_path):
    # copy the folder tree template
    shutil.copytree(FOLDER_TEMPLATE_PATH, proj_path)
        
def copy_proposal_sheet(proj_path, proj_name, cust_name, date):
    #create the proposal within the new project folder and rename it correctly
    quotes_path = os.path.join(proj_path, 'Quotes')
    shutil.copy(PROPOSAL_PATH, quotes_path )
    date_str = arrange_date(date)
    proj_name = '_'.join(proj_name.split(' '))
    proposal_str = '-'.join(['P', date_str, proj_name, cust_name, '(AL)'])
    prop_file = os.path.basename(PROPOSAL_PATH)
    prop_file_path = os.path.join(quotes_path, prop_file)
    new_prop_file_path = os.path.join(quotes_path, proposal_str)
    shutil.copy(PROPOSAL_PATH, quotes_path)
    shutil.move(prop_file_path, new_prop_file_path )
    return new_prop_file_path
    
        
    
def arrange_date(date_string):
    #rearrange the date string for the fiile name format
    date_params = date_string.split('/')
    date_params.reverse()
    new_date_str = ''.join(date_params)
    return new_date_str
    
    
def fill_prop_data(proposal_path, proj_name, cust_name, date, proj_address, sub_project_name, quote_num):
    # fill the data taken from the bid sheet into the proposal sheet
    proposal = openpyxl.load_workbook(proposal_path)
    summary = proposal.sheet('Summary')
    summary.cell('A1').value = proj_name
    summary.cell('A2').value = sub_project_name
    summary.cell('A3').value = proj_address
    summary.cell('G1').value = date
    summary.cell('G2').value = quote_num
    summary.cell('B8').value = cust_name
    proposal.save(proposal_path)
    
    

if __name__ == '__main__':
    quote_number, date, proj_name, sub_proj_name, proj_address, customer = get_bid_data()
    proj_path = create_new_bid_folder(proj_name, sub_proj_name)
    copy_folder_temp(proj_path)
    prop_path = copy_proposal_sheet(proj_path, proj_name, customer, date)
    fill_prop_data(prop_path, proj_name, customer, proj_address, date, sub_proj_name, quote_number)